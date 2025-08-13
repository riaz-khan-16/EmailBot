from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

class CrateExcel:
    def __init__(self,name):
        self.name=name
        
    def create(self):
        wb = Workbook()
        # Select the active worksheet
        ws = wb.active
        ws.title = "Sheet1"

        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 25
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Define center alignment
        center_alignment = Alignment(horizontal='center', vertical='center')
        # Add some data
        ws['A1'] = "Name"
        ws['A1'].fill=header_fill
        ws['A1'].alignment = center_alignment

        ws['B1'] = "Email"
        ws['B1'].fill=header_fill
        ws['B1'].alignment = center_alignment


        ws['C1'] = "Research Interest"
        ws['C1'].fill=header_fill
        ws['C1'].alignment = center_alignment

        ws['D1'] = "Website Link"
        ws['D1'].fill=header_fill
        ws['D1'].alignment = center_alignment


        # ws.append(["John Doe", "johndoe@example.com", "Control Systems"])
        
        # Save the workbook to a file
        wb.save('Storage/'+str(self.name)+".xlsx")

        print("Excel file created successfully!")


if __name__ == "__main__":
    obj=CrateExcel("test")
    obj.create()


    


        





